from os import listdir
from os.path import isfile, join
import pandas as pd
import openpyxl
from time import clock
import random
import math

class Utils:
    def __init__(self):
        self.ext_txt = ".txt"
        self.ext_dat = ".dat"
        self.ext_xlsx = ".xlsx"

    def read_cas9_scores_arr(self, path, f_num):
        tmp_str = ""
        with open(path + f_num + self.ext_txt, "r") as f:
            line_001 = f.readline().replace("\n", "")
            line_002 = f.readline().replace("\n", "")
            line_003 = f.readline().replace("\n", "")
            line_004 = f.readline().replace("\n", "")
            tmp_str = f.readline().replace("\n", "")
        print(f_num + " 's DeepCas9 score is read")
        return tmp_str.replace("(", "").replace(")", "").split(",")

    def read_excel_2_dataframe(self, path, f_name):
        return pd.read_excel(path + f_name + self.ext_xlsx)

    def read_dat_file_by_line_to_list(self,path):
        tmp_list = []
        with open(path + self.ext_dat, "r") as f:
            gene_str = ""
            CDC_str = ""
            idx = 0
            while True:
                tmp_line = f.readline().replace("\n", "")

                if tmp_line != '':
                    if "gene=ENSCSAG00000016985.1" in tmp_line:
                        break

                    if " gene " in tmp_line:
                        idx = idx + 1
                        print(str(idx) + ":::::::::::::::::::::::::::::::::::")

                    # print(tmp_line)
                    tmp_list.append(tmp_line)
                else:
                    break
        return tmp_list

    def make_txt_with_list(self, path, data_list):
        tmp_set = set()
        with open(path + "_" + str(clock()) + self.ext_txt, 'a') as f:
            for data_str in data_list:
                f.write(data_str + "\n")
                gene_id = data_str.split(" ")[-1]
                tmp_set.add(gene_id)
        return tmp_set

    def make_excel(self, result_dict, init, f_num):
        pam_len = len(init[0][0])
        add_1_len = init[1]
        spcr_len = init[2]
        add_2_len = init[3]
        clvg_after_pam = init[4]
        path = init[5]

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="INDEX")
        sheet.cell(row=row, column=2, value='Target gene name')
        sheet.cell(row=row, column=3, value='Description')
        sheet.cell(row=row, column=4, value='Ensembl transcript ID')
        sheet.cell(row=row, column=5, value='Ensembl Gene ID')
        sheet.cell(row=row, column=6, value='Position of Base After cut')
        sheet.cell(row=row, column=7, value='Strand')
        sheet.cell(row=row, column=8, value='sgRNA Target sequence')
        sheet.cell(row=row, column=9, value='Target context sequence')
        sheet.cell(row=row, column=10, value='PAM')
        sheet.cell(row=row, column=11, value='order sgRNA Target sequence')
        sheet.cell(row=row, column=12, value='order Target context sequence')
        sheet.cell(row=row, column=13, value='order PAM')
        sheet.cell(row=row, column=14, value='Exon Number')
        sheet.cell(row=row, column=15, value='DeepCas9 score')
        sheet.cell(row=row, column=16, value='target site (cleavage site)')

        for key, val_arr in result_dict.items():
            row = row + 1
            sheet.cell(row=row, column=1, value=(row - 1))
            if "ensembl".upper() != val_arr['Target gene name'].upper():
                sheet.cell(row=row, column=2, value=val_arr['Target gene name'])
            if 'Description' in val_arr:
                sheet.cell(row=row, column=3, value=val_arr['Description'])
            sheet.cell(row=row, column=4, value=val_arr['Ensembl transcript ID'])
            sheet.cell(row=row, column=5, value=val_arr['Ensembl Gene ID'])
            sheet.cell(row=row, column=6, value=val_arr['Position of Base After cut'])
            strand = val_arr['Strand']
            sheet.cell(row=row, column=7, value=strand)
            seq_str = val_arr['Target context sequence']
            if strand == '+':
                sheet.cell(row=row, column=8, value=seq_str[add_1_len:add_1_len + spcr_len])
                sheet.cell(row=row, column=9, value=seq_str)
                sheet.cell(row=row, column=10, value=seq_str[add_1_len + spcr_len:-add_2_len])
                sheet.cell(row=row, column=11, value=seq_str[add_1_len:add_1_len + spcr_len])
                sheet.cell(row=row, column=12, value=seq_str)
                sheet.cell(row=row, column=13, value=seq_str[add_1_len + spcr_len:-add_2_len])
            else:
                comp_seq_str = val_arr['Target context anti sequence']
                sheet.cell(row=row, column=8, value=seq_str[add_2_len + pam_len:add_2_len + pam_len + spcr_len])
                sheet.cell(row=row, column=9, value=seq_str)
                sheet.cell(row=row, column=10, value=seq_str[add_2_len:add_2_len + pam_len])
                sheet.cell(row=row, column=11, value=comp_seq_str[add_2_len + pam_len:add_2_len + pam_len + spcr_len][::-1])
                sheet.cell(row=row, column=12, value=comp_seq_str[::-1])
                sheet.cell(row=row, column=13, value=comp_seq_str[add_2_len:add_2_len + pam_len][::-1])
            sheet.cell(row=row, column=14, value=val_arr['Exon Number'])
            sheet.cell(row=row, column=16, value=val_arr['Ratio'])
        workbook.save(filename=path + "chlorochebus_sabaeus_" + f_num + "_" + str(clock()) + self.ext_xlsx)

    def make_excel_srt_by_deepcas9_scr(self, path, f_num, tmp_dict):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        for key, vals in tmp_dict.items():
            for tmp_list in sorted(vals, key=lambda tmp_list: tmp_list[0],reverse=True):
                col = 14
                for tmp_val in tmp_list:
                    if col >= 14:
                        sheet.cell(row=row, column=col, value=tmp_val)
                        if col >= 15:
                            col = 0
                    else:
                        if col < 3:
                            sheet.cell(row=row, column=col, value=tmp_val)
                        elif col == 3:
                            sheet.cell(row=row, column=col, value=key)
                            sheet.cell(row=row, column=col + 1, value=tmp_val)
                        else:
                            sheet.cell(row=row, column=col + 1, value=tmp_val)
                    col = col + 1
                row = row + 1

        # workbook.save(filename=path + f_num + "_" + str(clock()) + self.ext_xlsx)
        workbook.save(filename=path + f_num + self.ext_xlsx)

    def merge_deep_cas_9_data(self, tmp_arr, df_obj):
        tmp_dict = {'Ensembl transcript ID': [
            ['DeepCas9 score', 'target site (cleavage site)', 'Target gene name', 'Description', 'Ensembl Gene ID',
             'Position of Base After cut', 'Strand', 'sgRNA Target sequence', 'Target context sequence', 'PAM',
             'order sgRNA Target sequence', 'order Target context sequence', 'order PAM', 'Exon Number']]}
        data_obj = df_obj.to_dict()
        for i in range(len(tmp_arr)):
            # filtering out if target site (cleavage site) < 0.05, > 0.65
            if data_obj['target site (cleavage site)'][i] < 0.05:
                continue
            if data_obj['target site (cleavage site)'][i] > 0.65:
                continue
            trnscrpt_id = data_obj['Ensembl transcript ID'][i]
            tmp_list = []
            tmp_list.append(float(tmp_arr[i]))
            tmp_list.append(data_obj['target site (cleavage site)'][i])
            tmp_list.append(data_obj['Target gene name'][i])
            tmp_list.append(data_obj['Description'][i])
            tmp_list.append(data_obj['Ensembl Gene ID'][i])
            tmp_list.append(data_obj['Position of Base After cut'][i])
            tmp_list.append(data_obj['Strand'][i])
            tmp_list.append(data_obj['sgRNA Target sequence'][i])
            tmp_list.append(data_obj['Target context sequence'][i])
            tmp_list.append(data_obj['PAM'][i])
            tmp_list.append(data_obj['order sgRNA Target sequence'][i])
            tmp_list.append(data_obj['order Target context sequence'][i])
            tmp_list.append(data_obj['order PAM'][i])
            tmp_list.append(data_obj['Exon Number'][i])
            if trnscrpt_id not in tmp_dict.keys():
                tmp_dict[trnscrpt_id] = [tmp_list]
            else:
                tmp_dict[trnscrpt_id].append(tmp_list)
        return tmp_dict

    def read_txt_to_dict(self, init):
        path = init[0]
        file_ext = init[1]
        cnt = init[2]
        pam_len = len(init[3])

        result_dict = {}
        for idx in range(cnt):
            real_path = path + str(idx) + file_ext
            print(real_path)
            with open(real_path, "r") as f:
                header = f.readline()
                # print("header : " + header)# Sequence	0	1	2	3

                while True:
                    read_line = f.readline().replace("\n","")
                    if read_line == "":
                        break
                    read_line_arr = read_line.split("\t")
                    tmp_list = []
                    for i in range(1, len(read_line_arr)):
                        tmp_list.append(int(read_line_arr[i]))

                    spacer_seq = read_line_arr[0][:-pam_len]
                    if spacer_seq not in result_dict:
                        result_dict[spacer_seq] = tmp_list

        return result_dict

    """
    seq_cnt_dict = {trnscrpt_id: # of seq}
    """
    def make_tab_txt_seq_cnt_group_by_crpt_id(self, path, seq_cnt_dict):
        with open(path + str(clock()) + self.ext_txt, "w" ) as f:
            for crpt_id, seq_cnt in seq_cnt_dict.items():
                f.write(crpt_id + "\t" + str(seq_cnt) + "\n")

    """
    re_off_trgt_list = {transcript_id: [re_off_target_seq...]}
    """
    def make_tab_txt_re_off_target_seq(self, path, re_off_trgt_list):
        with open(path + str(clock()) + self.ext_txt, "w") as f:
            for crpt_id, seq_list in re_off_trgt_list.items():
                for seq_str in seq_list:
                    f.write(crpt_id + "\t" + seq_str + "\n")

    def make_excel_w_off_trgt(self, path, f_num, data_dict, seq_cnt_dict):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="INDEX")
        sheet.cell(row=row, column=2, value='Target gene name')
        sheet.cell(row=row, column=3, value='Description')
        sheet.cell(row=row, column=4, value='Ensembl transcript ID')
        sheet.cell(row=row, column=5, value='Ensembl Gene ID')
        sheet.cell(row=row, column=6, value='Position of Base After cut')
        sheet.cell(row=row, column=7, value='Strand')
        sheet.cell(row=row, column=8, value='sgRNA Target sequence')
        sheet.cell(row=row, column=9, value='Target context sequence')
        sheet.cell(row=row, column=10, value='PAM')
        sheet.cell(row=row, column=11, value='order sgRNA Target sequence')
        sheet.cell(row=row, column=12, value='order Target context sequence')
        sheet.cell(row=row, column=13, value='order PAM')
        sheet.cell(row=row, column=14, value='Exon Number')
        sheet.cell(row=row, column=15, value='DeepCas9 score')
        sheet.cell(row=row, column=16, value='target site (cleavage site)')

        # col_head = 17
        # to make columns for off_target results
        # for mis_match_cnt in range(len(off_trgt_dict[list(off_trgt_dict)[0]])):
        #     sheet.cell(row=row, column=col_head, value=str(mis_match_cnt))
        #     col_head += 1

        row += 1
        idx = 0

        for trnscrpt_id, vals_arr in data_dict.items():

            if trnscrpt_id not in seq_cnt_dict:
                seq_cnt_dict.update({trnscrpt_id: len(vals_arr)})

            for val_arr in vals_arr:
                idx += 1
                sheet.cell(row=row, column=1, value=str(idx))
                sheet.cell(row=row, column=2, value=val_arr[0])
                sheet.cell(row=row, column=3, value=val_arr[1])
                sheet.cell(row=row, column=4, value=trnscrpt_id)
                sheet.cell(row=row, column=5, value=val_arr[2])
                sheet.cell(row=row, column=6, value=val_arr[3])
                sheet.cell(row=row, column=7, value=val_arr[4])
                sheet.cell(row=row, column=8, value=val_arr[5])
                sheet.cell(row=row, column=9, value=val_arr[6])
                sheet.cell(row=row, column=10, value=val_arr[7])
                sheet.cell(row=row, column=11, value=val_arr[8])
                sheet.cell(row=row, column=12, value=val_arr[9])
                sheet.cell(row=row, column=13, value=val_arr[10])
                sheet.cell(row=row, column=14, value=val_arr[11])
                sheet.cell(row=row, column=15, value=val_arr[12])
                sheet.cell(row=row, column=16, value=val_arr[13])
                col = 17
                for result_cnt in val_arr[14]:
                    sheet.cell(row=row, column=col, value=str(result_cnt))
                    col += 1

                row += 1

        # workbook.save(filename=path + f_num + "_" + str(clock()) + self.ext_xlsx)
        workbook.save(filename=path + f_num + self.ext_xlsx)

        return seq_cnt_dict



